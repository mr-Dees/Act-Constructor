"""Извлечение текстового содержимого из файлов для контекста LLM.

Поддерживает: текст, PDF, Excel (.xlsx), Word (.docx).
Библиотеки опциональны — при отсутствии возвращается сообщение об ошибке.

Все парсеры (pypdf/openpyxl/python-docx) синхронные и CPU-bound. Прямой
вызов из async-функции блокирует event loop при больших файлах — на
SSE-стримах это приводит к замиранию всех соединений до окончания парсинга.
Используй ``extract_text_async`` из async-кода — он выполняет работу в
thread pool через ``asyncio.to_thread``.
"""

from __future__ import annotations

import asyncio
import io
import logging

logger = logging.getLogger("audit_workstation.domains.chat.file_extraction")

# Максимальный размер извлечённого текста (символов)
MAX_EXTRACTED_CHARS = 50_000


async def extract_text_async(
    file_data: bytes, mime_type: str, filename: str,
) -> str:
    """Async-обёртка над ``extract_text`` через ``asyncio.to_thread``.

    Используй из async-кода (SSE-стримы, polling, orchestrator), чтобы
    не блокировать event loop CPU-bound парсингом больших файлов.
    """
    return await asyncio.to_thread(extract_text, file_data, mime_type, filename)


def extract_text(file_data: bytes, mime_type: str, filename: str) -> str:
    """Извлекает текст из файла по MIME-типу. Синхронная, CPU-bound.

    Из async-кода зови ``extract_text_async`` — иначе блокируется event loop.

    Returns:
        Извлечённый текст или сообщение об ошибке.
    """
    try:
        if mime_type.startswith("text/"):
            return _extract_plain_text(file_data, filename)
        if mime_type == "application/pdf":
            return _extract_pdf(file_data, filename)
        if mime_type in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        ):
            return _extract_excel(file_data, filename)
        if mime_type in (
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        ):
            return _extract_docx(file_data, filename)
        if mime_type.startswith("image/"):
            return f"[Изображение: {filename}]"
        return f"[Файл: {filename} ({mime_type}) — извлечение текста не поддерживается]"
    except Exception as exc:
        logger.warning("Ошибка извлечения текста из %s: %s", filename, exc)
        return f"[Ошибка чтения файла {filename}: {exc}]"


def _truncate(text: str, filename: str) -> str:
    """Обрезает текст до лимита с пометкой."""
    if len(text) <= MAX_EXTRACTED_CHARS:
        return text
    return (
        text[:MAX_EXTRACTED_CHARS]
        + f"\n\n[... файл {filename} обрезан, показано {MAX_EXTRACTED_CHARS} символов]"
    )


def _extract_plain_text(file_data: bytes, filename: str) -> str:
    """Извлекает текст из текстовых файлов (UTF-8 / CP-1251)."""
    for encoding in ("utf-8", "cp1251", "latin-1"):
        try:
            text = file_data.decode(encoding)
            return _truncate(text, filename)
        except (UnicodeDecodeError, ValueError):
            continue
    return f"[Не удалось декодировать текстовый файл {filename}]"


def _extract_pdf(file_data: bytes, filename: str) -> str:
    """Извлекает текст из PDF."""
    try:
        import pypdf
    except ImportError:
        return f"[PDF файл {filename} — для чтения установите: pip install pypdf]"

    reader = pypdf.PdfReader(io.BytesIO(file_data))
    parts: list[str] = []
    total = 0
    for page in reader.pages:
        text = page.extract_text() or ""
        parts.append(text)
        total += len(text)
        if total > MAX_EXTRACTED_CHARS:
            break
    return _truncate("\n".join(parts), filename)


def _extract_excel(file_data: bytes, filename: str) -> str:
    """Извлекает данные из Excel (.xlsx)."""
    try:
        import openpyxl
    except ImportError:
        return f"[Excel файл {filename} — для чтения установите: pip install openpyxl]"

    wb = openpyxl.load_workbook(io.BytesIO(file_data), read_only=True, data_only=True)
    parts: list[str] = []
    total = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"--- Лист: {sheet_name} ---")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            line = "\t".join(cells)
            parts.append(line)
            total += len(line)
            if total > MAX_EXTRACTED_CHARS:
                break
        if total > MAX_EXTRACTED_CHARS:
            break
    wb.close()
    return _truncate("\n".join(parts), filename)


def _extract_docx(file_data: bytes, filename: str) -> str:
    """Извлекает текст из Word (.docx)."""
    try:
        import docx
    except ImportError:
        return f"[Word файл {filename} — для чтения установите: pip install python-docx]"

    doc = docx.Document(io.BytesIO(file_data))
    parts: list[str] = []
    total = 0
    for para in doc.paragraphs:
        text = para.text
        parts.append(text)
        total += len(text)
        if total > MAX_EXTRACTED_CHARS:
            break
    return _truncate("\n".join(parts), filename)
